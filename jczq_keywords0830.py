#!/usr/bin/env python3
"""
竞彩足球总进球数据抓取脚本（中文字段版）
提取字段：比赛日期、比赛时间、联赛、主队VS客队、比分
目标网站：500彩票网 (trade.500.com)
版本：2.8.0
修改：1.提示词01-06全部自动生成分析提示
     2.Excel列宽自适应调整，文字自动换行
     3.提示词03改为动态生成，内容与提示词02不同（侧重历史交锋和近期状态）
"""

import requests
import pandas as pd
from bs4 import BeautifulSoup
import json
import time
import random
import logging
from datetime import datetime, timedelta
import os
import sys
from typing import List, Dict, Optional

# 添加项目路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# 导入配置
try:
    from config import SCRAPER_CONFIG, HEADERS, OUTPUT_CONFIG

    BASE_URL = SCRAPER_CONFIG.get('base_url', 'https://trade.500.com/jczq/')
    TIMEOUT = SCRAPER_CONFIG.get('timeout', 15)
    RETRY_TIMES = SCRAPER_CONFIG.get('retry_times', 3)
    OUTPUT_DIR = OUTPUT_CONFIG.get('output_dir', 'data')
except ImportError:
    BASE_URL = 'https://trade.500.com/jczq/'
    TIMEOUT = 15
    RETRY_TIMES = 3
    OUTPUT_DIR = 'data'
    HEADERS = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br',
    }


# 设置日志
def setup_logging():
    """配置日志"""
    log_dir = OUTPUT_CONFIG.get('log_dir', 'logs') if 'OUTPUT_CONFIG' in globals() else 'logs'
    os.makedirs(log_dir, exist_ok=True)

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(os.path.join(log_dir, 'jczq_chinese_scraper.log'), encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)


logger = setup_logging()


class JczqChineseScraper:
    """竞彩足球中文字段数据抓取器"""

    def __init__(self, base_url: str = BASE_URL):
        self.base_url = base_url
        self.session = requests.Session()
        self.session.headers.update(HEADERS)
        logger.info(f"初始化中文字段抓取器，目标URL: {self.base_url}")

    def _random_delay(self, min_sec: float = 1.0, max_sec: float = 3.0):
        """随机延迟"""
        delay = random.uniform(min_sec, max_sec)
        time.sleep(delay)

    def fetch_html(self, date_str: str, retry: int = RETRY_TIMES) -> Optional[str]:
        """抓取HTML页面"""
        url = f"{self.base_url}?playid=270&g=2&date={date_str}"

        for attempt in range(retry):
            try:
                logger.info(f"抓取 {date_str}，尝试 {attempt + 1}/{retry}")
                response = self.session.get(url, timeout=TIMEOUT)
                response.raise_for_status()

                # 尝试多种编码
                for encoding in ['gb2312', 'gbk', 'utf-8']:
                    try:
                        response.encoding = encoding
                        html = response.text
                        if '<!doctype html' in html.lower():
                            logger.info(f"成功抓取，编码: {encoding}")
                            return html
                    except:
                        continue

                # 如果上面没成功，尝试直接解码
                html = response.content.decode('gb2312', errors='ignore')
                return html

            except requests.exceptions.RequestException as e:
                logger.warning(f"请求失败: {e}")
                if attempt < retry - 1:
                    time.sleep(2)

        logger.error(f"抓取 {date_str} 失败")
        return None

    def generate_prompt_01(self, league: str, matchup: str) -> str:
        """生成提示词01的自动分析提示（基本面）"""
        if not matchup or not league:
            return ""

        try:
            # 从主队VS客队字符串中提取主队和客队名称
            if ' VS ' in matchup:
                home_team, away_team = matchup.split(' VS ')
            elif ' vs ' in matchup:
                home_team, away_team = matchup.split(' vs ')
            else:
                home_team = away_team = ""

            if not home_team or not away_team:
                return ""

            # 生成分析提示
            prompt =  f"请根据参考指标自行搜索{home_team}VS{away_team}在{league}中的近期表现。"
            prompt += "两队参考指标：历史交锋、近期战绩、主场/客场表现、关键伤停、首发阵容/阵容深度，战术与战意。"
            prompt += "对本场比赛进行基本面分析,评估实力对比,判断优势方。"

            return prompt

        except Exception as e:
            logger.error(f"生成提示词01时出错: {e}")
            return ""

    def generate_prompt_02(self, league: str, home_team: str, away_team: str) -> str:
        """生成提示词02的自动分析提示（定位赔率区间，不含实际赔率）"""
        if not home_team or not away_team or not league:
            return "在意甲中,基本面:客队占优.以初盘3.00/3.00/2.40为准,定位赔率区间和赔率位置."
        return f"在{league}中,本场比赛{home_team}VS{away_team}基本面:客队占优.以威廉希尔初盘3.00/3.00/2.40为准,定位赔率区间和赔率位置."

    def generate_prompt_03(self, league: str, home_team: str, away_team: str) -> str:
        """
        生成提示词03的自动分析提示（侧重历史交锋和近期状态，与提示词02不同）
        不包含任何赔率信息
        """
        if not home_team or not away_team or not league:
            return "请分析本场比赛双方的历史交锋记录和近期状态，评估实力对比。"
        return f"在{league}中{home_team}与{away_team}基本面:威廉初盘赔率为:,终盘赔率为:.请分析本次比赛真实意图和诱盘"

    def generate_prompt_04(self) -> str:
        """生成提示词04的自动分析提示"""
        return "请根据上面生成的交叉验证分析报告,给出本次比赛的最佳投注策略,需要注明投注额的比例分配"

    def generate_prompt_05(self) -> str:
        """生成提示词05的自动分析提示"""
        return "请对上面的投注策略给出盈利分析"

    def generate_prompt_06(self) -> str:
        """生成提示词06的自动分析提示"""
        return "请结合上面的基本面分析，以及附件中竞彩官方欧赔指数和让球指数数据，对本场比赛的竞彩官方数据提供细化分析"

    def parse_match_row(self, row) -> Dict:
        """解析比赛行 - 使用中文字段名，合并主队和客队，添加比赛时间"""
        try:
            # 提取联赛名称，将"沙特职业联赛"改为"沙特联"
            league_name = row.get('data-simpleleague', '')
            if '沙特职业联赛' in league_name:
                league_name = league_name.replace('沙特职业联赛', '沙特联')

            # 提取主队和客队名称
            home_team = row.get('data-homesxname', '')
            away_team = row.get('data-awaysxname', '')

            # 合并主队和客队为"主队 VS 客队"格式
            team_matchup = f"{home_team} VS {away_team}" if home_team and away_team else ""

            # 提取比赛时间 - 从data-matchtime属性获取
            match_time_full = row.get('data-matchtime', '')
            match_time = ''

            if match_time_full:
                try:
                    # 尝试解析完整时间格式，提取HH:MM部分
                    if ':' in match_time_full:
                        time_parts = match_time_full.split(':')
                        if len(time_parts) >= 2:
                            hour = time_parts[0].zfill(2)
                            minute = time_parts[1].zfill(2)
                            match_time = f"{hour}:{minute}"
                except:
                    match_time = ''

            # 提取基础字段，使用中文名称
            match_data = {
                '比赛日期': row.get('data-matchdate', ''),
                '比赛时间': match_time,
                '联赛': league_name,
                '主队VS客队': team_matchup,
            }

            # 提取比分
            team_div = row.select_one('.team')
            if team_div:
                score_elem = team_div.select_one('.score')
                if score_elem:
                    match_data['比分'] = score_elem.text.strip()
                else:
                    bf_elem = team_div.select_one('.team-bf')
                    if bf_elem:
                        match_data['比分'] = bf_elem.text.strip()
                    else:
                        match_data['比分'] = ''
            else:
                match_data['比分'] = ''

            if '比分' not in match_data or not match_data['比分']:
                match_data['比分'] = ''

            # 添加提示词列 - 全部自动生成
            match_data['提示词01'] = self.generate_prompt_01(league_name, team_matchup)
            match_data['提示词02'] = self.generate_prompt_02(league_name, home_team, away_team)
            match_data['提示词03'] = self.generate_prompt_03(league_name, home_team, away_team)   # 动态生成，与02不同
            match_data['提示词04'] = self.generate_prompt_04()
            match_data['提示词05'] = self.generate_prompt_05()
            match_data['提示词06'] = self.generate_prompt_06()

            return match_data

        except Exception as e:
            logger.error(f"解析行失败: {e}")
            return {
                '比赛日期': '',
                '比赛时间': '',
                '联赛': '',
                '主队VS客队': '',
                '比分': '',
                '提示词01': '',
                '提示词02': '',
                '提示词03': '',
                '提示词04': '',
                '提示词05': '',
                '提示词06': '',
                '错误信息': str(e)
            }

    def parse_html(self, html: str, date_str: str) -> List[Dict]:
        """解析HTML - 使用中文字段名"""
        if not html:
            return []

        try:
            soup = BeautifulSoup(html, 'html.parser')
            rows = soup.select('tr.bet-tb-tr[data-fixtureid]')
            logger.info(f"找到 {len(rows)} 场比赛")

            matches = []
            for row in rows:
                match_data = self.parse_match_row(row)
                matches.append(match_data)

            return matches

        except Exception as e:
            logger.error(f"解析HTML失败: {e}")
            return []

    def scrape_date(self, date_str: str) -> List[Dict]:
        """抓取单日数据"""
        html = self.fetch_html(date_str)
        if html:
            return self.parse_html(html, date_str)
        return []

    def scrape_date_range(self, start_date: str, end_date: str) -> Dict[str, List[Dict]]:
        """抓取日期范围数据"""
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')

        all_data = {}
        current = start

        while current <= end:
            date_str = current.strftime('%Y-%m-%d')
            logger.info(f"处理 {date_str}")

            matches = self.scrape_date(date_str)
            if matches:
                all_data[date_str] = matches
                logger.info(f"成功抓取 {len(matches)} 场比赛")

            current += timedelta(days=1)
            if current <= end:
                self._random_delay(1, 2)

        return all_data

    def save_to_json(self, data: List[Dict], filename: str):
        """保存为JSON"""
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        filepath = os.path.join(OUTPUT_DIR, filename)

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        logger.info(f"JSON保存到: {filepath}")

    def save_to_csv(self, matches: List[Dict], filename: str):
        """保存为CSV - 使用中文列名，包含提示词列"""
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        filepath = os.path.join(OUTPUT_DIR, filename)

        if matches:
            df = pd.DataFrame(matches)
            columns_order = [
                '比赛日期', '比赛时间', '联赛', '主队VS客队', '比分',
                '提示词01', '提示词02', '提示词03', '提示词04', '提示词05', '提示词06'
            ]
            existing_columns = [col for col in columns_order if col in df.columns]
            df = df[existing_columns]
            df.to_csv(filepath, index=False, encoding='utf-8-sig')
            logger.info(f"CSV保存到: {filepath}")

    def save_to_excel(self, matches: List[Dict], filename: str):
        """保存为Excel - 自适应列宽，文字自动换行"""
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        filepath = os.path.join(OUTPUT_DIR, filename)

        if not matches:
            logger.warning("没有数据可保存")
            return False

        try:
            df = pd.DataFrame(matches)

            columns_order = [
                '比赛日期', '比赛时间', '联赛', '主队VS客队', '比分',
                '提示词01', '提示词02', '提示词03', '提示词04', '提示词05', '提示词06'
            ]
            existing_columns = [col for col in columns_order if col in df.columns]
            df = df[existing_columns]

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='比赛信息')

                workbook = writer.book
                worksheet = writer.sheets['比赛信息']

                self._adjust_column_widths(worksheet, df)

                # 设置自动换行样式
                wrap_alignment = 'wrap_alignment'
                if wrap_alignment not in workbook.named_styles:
                    from openpyxl.styles import Alignment, NamedStyle
                    wrap_style = NamedStyle(name=wrap_alignment)
                    wrap_style.alignment = Alignment(wrap_text=True, vertical='top')
                    workbook.add_named_style(wrap_style)

                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1,
                                               max_col=len(existing_columns)):
                    for cell in row:
                        cell.style = wrap_alignment

                from openpyxl.styles import Alignment, Font
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                header_font = Font(bold=True)

                for cell in worksheet[1]:
                    cell.alignment = header_alignment
                    cell.font = header_font

            logger.info(f"Excel保存到: {filepath}")
            logger.info("已应用：自适应列宽、文字自动换行、标题行居中加粗")

            if len(df) > 0 and '提示词01' in df.columns:
                for i in range(1, 7):
                    col_name = f'提示词{i:02d}'
                    if col_name in df.columns:
                        prompt_example = df.iloc[0][col_name]
                        if prompt_example:
                            logger.info(f"{col_name}示例: {prompt_example[:50]}...")

            return True

        except Exception as e:
            logger.error(f"保存Excel时出错: {e}")
            try:
                if matches:
                    df = pd.DataFrame(matches)
                    columns_order = [
                        '比赛日期', '比赛时间', '联赛', '主队VS客队', '比分',
                        '提示词01', '提示词02', '提示词03', '提示词04', '提示词05', '提示词06'
                    ]
                    existing_columns = [col for col in columns_order if col in df.columns]
                    df = df[existing_columns]
                    df.to_excel(filepath, index=False, sheet_name='比赛信息')
                    logger.info(f"Excel保存到: {filepath} (普通版)")
                    return True
            except Exception as e2:
                logger.error(f"保存普通版Excel也失败: {e2}")
                return False

    def _adjust_column_widths(self, worksheet, df):
        """调整列宽自适应"""
        try:
            from openpyxl.utils import get_column_letter

            base_widths = {
                '比赛日期': 12,
                '比赛时间': 10,
                '联赛': 20,
                '主队VS客队': 25,
                '比分': 8,
                '提示词01': 40,
                '提示词02': 40,
                '提示词03': 50,
                '提示词04': 35,
                '提示词05': 25,
                '提示词06': 45,
            }

            for i, column in enumerate(df.columns, 1):
                column_letter = get_column_letter(i)

                column_name_width = len(str(column)) * 1.2

                max_content_width = 0
                for cell_value in df[column]:
                    if pd.notna(cell_value):
                        lines = str(cell_value).split('\n')
                        if lines:
                            line_width = max(len(line) for line in lines) * 1.1
                            max_content_width = max(max_content_width, line_width)

                base_width = base_widths.get(column, 15)
                max_width = max(column_name_width, max_content_width, base_width)
                max_width = min(max_width, 80)

                worksheet.column_dimensions[column_letter].width = max_width

            logger.info("已自动调整列宽")

        except Exception as e:
            logger.warning(f"调整列宽时出错，使用默认设置: {e}")
            default_widths = {
                'A': 12, 'B': 10, 'C': 20, 'D': 25, 'E': 8,
                'F': 40, 'G': 40, 'H': 50, 'I': 35, 'J': 25, 'K': 45,
            }
            for col_letter, width in default_widths.items():
                if col_letter in worksheet.column_dimensions:
                    worksheet.column_dimensions[col_letter].width = width

    @staticmethod
    def merge_excel_files():
        """合并Excel文件 - 自适应列宽版"""
        if not os.path.exists(OUTPUT_DIR):
            print(f"❌ 目录 {OUTPUT_DIR} 不存在")
            return

        excel_files = [f for f in os.listdir(OUTPUT_DIR) if f.endswith('.xlsx') or f.endswith('.xls')]

        if not excel_files:
            print(f"❌ 在 {OUTPUT_DIR} 目录中未找到Excel文件")
            return

        print(f"\n在 {OUTPUT_DIR} 目录中找到 {len(excel_files)} 个Excel文件:")
        print("-" * 60)
        for i, file in enumerate(excel_files, 1):
            print(f"{i:2d}. {file}")
        print("-" * 60)

        file_nums = input("请输入要合并的文件序号（用逗号分隔，如 1,2,3 或输入 all 合并全部）: ").strip()

        selected_files = []
        if file_nums.lower() == 'all':
            selected_files = excel_files
        else:
            try:
                indices = [int(num.strip()) - 1 for num in file_nums.split(',')]
                for idx in indices:
                    if 0 <= idx < len(excel_files):
                        selected_files.append(excel_files[idx])
                    else:
                        print(f"❌ 序号 {idx + 1} 无效，跳过")
            except ValueError:
                print("❌ 输入格式错误")
                return

        if not selected_files:
            print("❌ 未选择任何文件")
            return

        print(f"\n将要合并以下 {len(selected_files)} 个文件:")
        for file in selected_files:
            print(f"  - {file}")

        confirm = input("\n是否确认合并这些文件? (y/n): ").lower()
        if confirm != 'y':
            print("❌ 已取消合并操作")
            return

        all_data = []
        processed_count = 0

        for file in selected_files:
            filepath = os.path.join(OUTPUT_DIR, file)
            try:
                df = pd.read_excel(filepath)
                if not df.empty:
                    all_data.append(df)
                    processed_count += 1
                    print(f"✅ 已加载: {file} ({len(df)} 行)")
                else:
                    print(f"⚠️  文件为空: {file}")
            except Exception as e:
                print(f"❌ 读取文件失败 {file}: {e}")

        if not all_data:
            print("❌ 没有有效数据可合并")
            return

        merged_df = pd.concat(all_data, ignore_index=True)

        expected_columns = [
            '比赛日期', '比赛时间', '联赛', '主队VS客队', '比分',
            '提示词01', '提示词02', '提示词03', '提示词04', '提示词05', '提示词06'
        ]

        for col in expected_columns:
            if col not in merged_df.columns:
                merged_df[col] = ''

        merged_df = merged_df[expected_columns]

        before_count = len(merged_df)
        merged_df = merged_df.drop_duplicates()
        after_count = len(merged_df)

        if before_count != after_count:
            print(f"⚠️  已移除 {before_count - after_count} 条重复记录")

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"合并足球数据_{timestamp}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, output_filename)

        try:
            scraper = JczqChineseScraper()
            success = scraper._save_merged_excel(merged_df, output_path, processed_count)

            if success:
                print(f"\n✅ 合并完成!")
                print(f"   合并了 {processed_count} 个文件")
                print(f"   总数据行数: {len(merged_df)}")
                print(f"   保存为: {output_filename} (已应用自适应列宽和自动换行)")

                print(f"\n表格包含 {len(merged_df.columns)} 列:")
                for i, col in enumerate(merged_df.columns, 1):
                    print(f"  {i:2d}. {col}")

                print(f"\n提示词填充情况:")
                for i in range(1, 7):
                    col_name = f'提示词{i:02d}'
                    if col_name in merged_df.columns:
                        non_empty = merged_df[col_name].notna().sum()
                        print(f"  {col_name}: {non_empty}条非空")

                print("\n合并后数据统计:")
                print("-" * 40)
                if '比赛日期' in merged_df.columns:
                    print(f"比赛日期范围: {merged_df['比赛日期'].min()} 至 {merged_df['比赛日期'].max()}")
                if '联赛' in merged_df.columns:
                    print(f"联赛数量: {merged_df['联赛'].nunique()}")
                print(f"比赛总数: {len(merged_df)}")

            else:
                merged_df.to_excel(output_path, index=False, sheet_name='合并数据')
                print(f"\n✅ 合并完成（普通版）!")
                print(f"   合并了 {processed_count} 个文件")
                print(f"   总数据行数: {len(merged_df)}")
                print(f"   保存为: {output_filename}")

        except Exception as e:
            print(f"❌ 合并Excel时出错: {e}")
            try:
                merged_df.to_excel(output_path, index=False, sheet_name='合并数据')
                print(f"\n✅ 合并完成（普通版）!")
                print(f"   保存为: {output_filename}")
            except Exception as e2:
                print(f"❌ 保存普通版也失败: {e2}")

    def _save_merged_excel(self, merged_df, output_path, processed_count):
        """保存合并的Excel文件，应用自适应列宽"""
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                merged_df.to_excel(writer, index=False, sheet_name='合并数据')

                workbook = writer.book
                worksheet = writer.sheets['合并数据']

                self._adjust_column_widths(worksheet, merged_df)

                from openpyxl.styles import Alignment, NamedStyle
                wrap_alignment = 'wrap_alignment_merged'
                if wrap_alignment not in workbook.named_styles:
                    wrap_style = NamedStyle(name=wrap_alignment)
                    wrap_style.alignment = Alignment(wrap_text=True, vertical='top')
                    workbook.add_named_style(wrap_style)

                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1,
                                               max_col=len(merged_df.columns)):
                    for cell in row:
                        cell.style = wrap_alignment

                from openpyxl.styles import Font
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                header_font = Font(bold=True)

                for cell in worksheet[1]:
                    cell.alignment = header_alignment
                    cell.font = header_font

                stat_row = len(merged_df) + 3
                worksheet.cell(row=stat_row, column=1, value="合并统计信息:")
                worksheet.cell(row=stat_row, column=1).font = Font(bold=True)

                stat_row += 1
                worksheet.cell(row=stat_row, column=1, value="合并文件数:")
                worksheet.cell(row=stat_row, column=2, value=processed_count)

                stat_row += 1
                worksheet.cell(row=stat_row, column=1, value="比赛总数:")
                worksheet.cell(row=stat_row, column=2, value=len(merged_df))

                stat_row += 1
                worksheet.cell(row=stat_row, column=1, value="合并时间:")
                worksheet.cell(row=stat_row, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

            print(f"   已应用：自适应列宽、文字自动换行、标题行居中加粗")
            return True

        except Exception as e:
            print(f"⚠️ 应用自适应列宽时出错: {e}")
            return False

    def display_matches(self, matches: List[Dict], title: str = "比赛列表"):
        """显示比赛列表"""
        if not matches:
            print("暂无比赛数据")
            return

        print(f"\n{title}:")
        print("=" * 120)
        print(
            f"{'序号':<4} {'比赛日期':<12} {'时间':<8} {'联赛':<15} {'主队VS客队':<35} {'比分':<8} {'提示词状态':<12}")
        print("=" * 120)

        for i, match in enumerate(matches, 1):
            team_matchup = match.get('主队VS客队', '')
            display_teams = team_matchup[:33] + '...' if len(team_matchup) > 35 else team_matchup

            prompt_count = sum(1 for j in range(1, 7) if match.get(f'提示词{j:02d}', ''))
            prompt_status = f"{prompt_count}/6已生成"

            print(f"{i:<4} "
                  f"{match.get('比赛日期', '')[:12]:<12} "
                  f"{match.get('比赛时间', ''):<8} "
                  f"{match.get('联赛', '')[:15]:<15} "
                  f"{display_teams:<35} "
                  f"{match.get('比分', ''):<8} "
                  f"{prompt_status:<12}")

        if matches and any(match.get('提示词01', '') for match in matches):
            print(f"\n提示词内容示例:")
            match = matches[0]
            for i in range(1, 7):
                col_name = f'提示词{i:02d}'
                prompt = match.get(col_name, '')
                if prompt:
                    print(f"  {col_name}: {prompt[:70]}...")

    def get_today_matches_and_save(self):
        """获取今日赛事并自动保存"""
        today = datetime.now().strftime('%Y-%m-%d')
        print(f"正在获取今日 ({today}) 的赛事数据...")

        matches = self.scrape_date(today)
        if matches:
            print(f"✅ 找到 {len(matches)} 场比赛")
            self.display_matches(matches, f"今日比赛 ({today})")

            if matches:
                total_prompts = 0
                for i in range(1, 7):
                    col_name = f'提示词{i:02d}'
                    generated = sum(1 for match in matches if match.get(col_name, ''))
                    total_prompts += generated
                    print(f"  {col_name}: {generated}场已生成")

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            self.save_to_json(matches, f"今日足球数据_{today}_{timestamp}.json")
            self.save_to_csv(matches, f"今日足球数据_{today}_{timestamp}.csv")
            success = self.save_to_excel(matches, f"今日足球数据_{today}_{timestamp}.xlsx")

            if success:
                print(f"\n✅ 今日赛事数据已自动保存到 data/ 目录")
                print(f"   保存的文件：今日足球数据_{today}_{timestamp}.*")
                print(f"   表格包含列：比赛日期、比赛时间、联赛、主队VS客队、比分、提示词01-06")
                print(f"   提示词03已动态生成，内容与提示词02不同")
                print(f"   已应用：自适应列宽、文字自动换行、标题行居中加粗")
            else:
                print("⚠️  Excel保存失败，但JSON和CSV已保存")

            if matches:
                time_stats = {}
                for match in matches:
                    match_time = match.get('比赛时间', '未知')
                    time_stats[match_time] = time_stats.get(match_time, 0) + 1

                print(f"\n开球时间分布:")
                for time_str, count in sorted(time_stats.items()):
                    if time_str:
                        print(f"  {time_str}: {count}场")
        else:
            print("❌ 今日无比赛")


def parse_date_input(date_str: str, default_year: str = "2006") -> str:
    """解析日期输入，自动补全年份"""
    if not date_str:
        return f"{default_year}-01-01"

    date_str = date_str.strip()

    if len(date_str) == 10 and '-' in date_str:
        try:
            datetime.strptime(date_str, '%Y-%m-%d')
            return date_str
        except ValueError:
            pass

    if len(date_str) == 5 and '-' in date_str:
        try:
            month_day = date_str.split('-')
            if len(month_day) == 2:
                month, day = month_day
                test_date = f"{default_year}-{month.zfill(2)}-{day.zfill(2)}"
                datetime.strptime(test_date, '%Y-%m-%d')
                return test_date
        except ValueError:
            pass

    if '-' in date_str and len(date_str) <= 5:
        try:
            parts = date_str.split('-')
            if len(parts) == 2:
                month, day = parts
                month = month.zfill(2)
                day = day.zfill(2)
                test_date = f"{default_year}-{month}-{day}"
                datetime.strptime(test_date, '%Y-%m-%d')
                return test_date
        except ValueError:
            pass

    try:
        datetime.strptime(date_str, '%Y-%m-%d')
        return date_str
    except ValueError:
        raise ValueError(f"无法解析日期: {date_str}")


def main():
    """主函数"""
    current_year = datetime.now().strftime('%Y')
    default_year = "2006"

    print("=" * 70)
    print("竞彩足球中文字段数据抓取工具 v2.8.0")
    print(f"当前年份：{current_year} (输入日期时年份默认为{default_year})")
    print("提取字段：比赛日期、比赛时间、联赛、主队VS客队、比分")
    print("提示词01-06全部自动生成（提示词03动态生成，内容与提示词02不同）")
    print("Excel优化：自适应列宽、文字自动换行、标题行居中加粗")
    print("联赛处理：沙特职业联赛 → 沙特联")
    print("=" * 70)

    scraper = JczqChineseScraper()

    while True:
        print("\n请选择操作:")
        print("1. 抓取单日数据")
        print("2. 抓取日期范围数据")
        print("3. 获取今日赛事并保存")
        print("4. 合并Excel表格（自适应列宽版）")
        print("5. 显示字段说明")
        print("6. 退出")

        choice = input("请输入选项 (1-6): ").strip()

        if choice == '1':
            date_prompt = f"请输入日期 (格式: MM-DD，年份默认为{default_year})，如 01-01: "
            date_input = input(date_prompt).strip()

            try:
                date_str = parse_date_input(date_input, default_year)
                print(f"解析后的日期: {date_str}")
                print(f"正在抓取 {date_str} 的数据...")

                matches = scraper.scrape_date(date_str)
                if matches:
                    print(f"✅ 成功抓取 {len(matches)} 场比赛")
                    scraper.display_matches(matches, f"{date_str} 比赛列表")

                    if matches:
                        print(f"\n提示词自动生成情况:")
                        for i in range(1, 7):
                            col_name = f'提示词{i:02d}'
                            generated = sum(1 for match in matches if match.get(col_name, ''))
                            print(f"  {col_name}: {generated}场已生成")

                    if matches:
                        print(f"\n表格将包含以下列（自适应列宽）:")
                        columns = [
                            '比赛日期', '比赛时间', '联赛', '主队VS客队', '比分',
                            '提示词01', '提示词02', '提示词03', '提示词04', '提示词05', '提示词06'
                        ]
                        for i, col in enumerate(columns, 1):
                            print(f"  {i:2d}. {col}")

                    time_stats = {}
                    for match in matches:
                        match_time = match.get('比赛时间', '未知')
                        time_stats[match_time] = time_stats.get(match_time, 0) + 1

                    print(f"\n开球时间分布:")
                    for time_str, count in sorted(time_stats.items()):
                        if time_str:
                            print(f"  {time_str}: {count}场")

                    save = input("\n是否保存数据? (y/n): ").lower()
                    if save == 'y':
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        scraper.save_to_json(matches, f"足球数据_{date_str}_{timestamp}.json")
                        scraper.save_to_csv(matches, f"足球数据_{date_str}_{timestamp}.csv")
                        success = scraper.save_to_excel(matches, f"足球数据_{date_str}_{timestamp}.xlsx")

                        if success:
                            print("✅ 数据已保存到 data/ 目录")
                            print(f"   表格包含 {len(matches[0].keys())} 列")
                            print(f"   提示词03已动态生成，内容与提示词02不同")
                            print(f"   已应用：自适应列宽、文字自动换行、标题行居中加粗")
                        else:
                            print("⚠️  Excel保存失败，但JSON和CSV已保存")
                else:
                    print("❌ 未找到数据")

            except ValueError as e:
                print(f"❌ 日期格式错误: {e}")
                print(f"请输入有效的日期格式，如: 01-15 或 1-15")

        elif choice == '2':
            start_prompt = f"开始日期 (格式: MM-DD，年份默认为{default_year}): "
            end_prompt = f"结束日期 (格式: MM-DD，年份默认为{default_year}): "

            start_input = input(start_prompt).strip()
            end_input = input(end_prompt).strip()

            try:
                start = parse_date_input(start_input, default_year)
                end = parse_date_input(end_input, default_year)

                start_dt = datetime.strptime(start, '%Y-%m-%d')
                end_dt = datetime.strptime(end, '%Y-%m-%d')

                if start_dt > end_dt:
                    print("❌ 开始日期不能晚于结束日期")
                    continue

                print(f"解析后的日期范围: {start} 到 {end}")
                print(f"正在抓取 {start} 到 {end} 的数据...")
                all_data = scraper.scrape_date_range(start, end)

                total = sum(len(matches) for matches in all_data.values())
                print(f"✅ 完成! 总比赛数: {total}")

                if all_data:
                    all_matches = []
                    for date_matches in all_data.values():
                        all_matches.extend(date_matches)

                    print("\n数据汇总:")
                    print(f"日期范围: {start} 至 {end}")
                    print(f"比赛总数: {total}")
                    print(f"表格列数: 11列 (包含6个提示词列)")
                    print(f"表格优化: 自适应列宽、文字自动换行")

                    print(f"\n提示词自动生成情况:")
                    for i in range(1, 7):
                        col_name = f'提示词{i:02d}'
                        generated = sum(1 for match in all_matches if match.get(col_name, ''))
                        print(f"  {col_name}: {generated}场已生成")

                    time_stats = {}
                    for match in all_matches:
                        match_time = match.get('比赛时间', '未知')
                        time_stats[match_time] = time_stats.get(match_time, 0) + 1

                    print(f"\n开球时间分布（前10）:")
                    sorted_times = sorted([(t, c) for t, c in time_stats.items() if t],
                                          key=lambda x: x[1], reverse=True)
                    for time_str, count in sorted_times[:10]:
                        print(f"  {time_str}: {count}场")

                    league_stats = {}
                    saudi_count = 0
                    for match in all_matches:
                        league = match.get('联赛', '未知联赛')
                        league_stats[league] = league_stats.get(league, 0) + 1
                        if '沙特联' in league:
                            saudi_count += 1

                    print(f"\n联赛分布（前10）:")
                    sorted_leagues = sorted(league_stats.items(), key=lambda x: x[1], reverse=True)
                    for league, count in sorted_leagues[:10]:
                        print(f"  {league}: {count}场")

                    if saudi_count > 0:
                        print(f"\n注：已将 {saudi_count} 场'沙特职业联赛'转换为'沙特联'")

                    save = input("\n是否保存数据? (y/n): ").lower()
                    if save == 'y':
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        scraper.save_to_json(all_matches, f"足球数据_{start}_至_{end}_{timestamp}.json")
                        scraper.save_to_csv(all_matches, f"足球数据_{start}_至_{end}_{timestamp}.csv")
                        success = scraper.save_to_excel(all_matches, f"足球数据_{start}_至_{end}_{timestamp}.xlsx")

                        if success:
                            print("✅ 数据已保存到 data/ 目录")
                            print(f"   表格包含11列：比赛基本信息 + 提示词01-06列")
                            print(f"   提示词03已动态生成，内容与提示词02不同")
                            print(f"   已应用：自适应列宽、文字自动换行、标题行居中加粗")
                        else:
                            print("⚠️  Excel保存失败，但JSON和CSV已保存")

            except ValueError as e:
                print(f"❌ 日期格式错误: {e}")
                print(f"请输入有效的日期格式，如: 01-15 或 1-15")

        elif choice == '3':
            scraper.get_today_matches_and_save()

        elif choice == '4':
            JczqChineseScraper.merge_excel_files()

        elif choice == '5':
            print("\n字段说明:")
            print("=" * 65)
            print("基础字段:")
            print("  比赛日期: 比赛进行的日期 (YYYY-MM-DD)")
            print("  比赛时间: 比赛开球时间 (HH:MM格式，如 19:30)")
            print("  联赛:     比赛所属的联赛名称")
            print("          注：'沙特职业联赛'会自动转换为'沙特联'")
            print("  主队VS客队: 主队和客队名称合并，格式为'主队 VS 客队'")
            print("  比分:     比赛最终比分（如3:2）")
            print("\n新增字段（全部自动生成）:")
            print("  提示词01: 基本面分析提示（动态生成）")
            print("  提示词02: 定位赔率区间提示（动态生成，不含实际赔率）")
            print("  提示词03: 历史交锋和近期状态分析提示（动态生成，与02不同）")
            print("  提示词04: 最佳投注策略")
            print("  提示词05: 盈利分析")
            print("  提示词06: 竞彩官方数据分析")
            print("\nExcel优化功能:")
            print("  - 自适应列宽: 根据内容自动调整列宽")
            print("  - 文字自动换行: 长文本自动换行显示")
            print("  - 标题行居中加粗: 标题行美观显示")
            print("  - 包含11列数据：5个基础列 + 6个提示词列")
            print("  - 所有提示词列均已自动填充")
            print("\n日期输入说明:")
            print("  - 单日数据: 输入 MM-DD (如: 01-15)")
            print("  - 日期范围: 分别输入开始和结束的 MM-DD")
            print(f"  - 年份自动设置为: {default_year}")
            print("=" * 65)

        elif choice == '6':
            print("感谢使用，再见!")
            break

        else:
            print("❌ 无效选项")


if __name__ == "__main__":
    main()